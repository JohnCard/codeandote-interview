from openpyxl.utils import get_column_letter, range_boundaries
from consuming_api.models import Country
import requests
import matplotlib
matplotlib.use('Agg')  # ⚠️ Importante para evitar errores en macOS
import matplotlib.pyplot as plt
import numpy as np
import io
from openpyxl.drawing.image import Image as OpenpyxlImage
from .styles import BORDER
import random
from faker import Faker

fake = Faker()

img_list = [
    'https://fakestoreapi.com/img/81fPKd-2AYL._AC_SL1500_t.png', 'https://fakestoreapi.com/img/71-3HjGNDUL._AC_SY879._SX._UX._SY._UY_t.png', 'https://fakestoreapi.com/img/71li-ujtlUL._AC_UX679_t.png', 'https://fakestoreapi.com/img/71YXzeOuslL._AC_UY879_t.png', 'https://fakestoreapi.com/img/71pWzhdJNwL._AC_UL640_QL65_ML3_t.png', 'https://fakestoreapi.com/img/61sbMiUnoGL._AC_UL640_QL65_ML3_t.png', 'https://fakestoreapi.com/img/71YAIFU48IL._AC_UL640_QL65_ML3_t.png', 'https://fakestoreapi.com/img/51UDEzMJVpL._AC_UL640_QL65_ML3_t.png', 'https://fakestoreapi.com/img/61IBBVJvSDL._AC_SY879_t.png', 'https://fakestoreapi.com/img/61U7T1koQqL._AC_SX679_t.png', 'https://fakestoreapi.com/img/71kWymZ+c+L._AC_SX679_t.png', 'https://fakestoreapi.com/img/61mtL65D4cL._AC_SX679_t.png', 'https://fakestoreapi.com/img/81QpkIctqPL._AC_SX679_t.png', 'https://fakestoreapi.com/img/81Zt42ioCgL._AC_SX679_t.png', 'https://fakestoreapi.com/img/51Y5NI-I5jL._AC_UX679_t.png', 'https://fakestoreapi.com/img/81XH0e8fefL._AC_UY879_t.png', 'https://fakestoreapi.com/img/71HblAHs5xL._AC_UY879_-2t.png', 'https://fakestoreapi.com/img/71z3kpMAYsL._AC_UY879_t.png', 'https://fakestoreapi.com/img/51eg55uWmdL._AC_UX679_t.png', 'https://fakestoreapi.com/img/61pHAEJ4NML._AC_UX679_t.png']

# Additional functions

def generate_product_name():
    brands = ['NovaTech', 'EcoStyle', 'UrbanHome', 'QuickWear', 'SmartLife', 'ZenGo']
    product_types = [
        'Chair', 'Laptop', 'Tablet', 'Sofa', 'T-Shirt', 'Blender', 'Microwave',
        'Smartwatch', 'TV', 'Backpack', 'Sneakers', 'Desk Lamp', 'Vacuum', 'Camera'
    ]

    brand = random.choice(brands)
    adjective = fake.word().capitalize()
    product = random.choice(product_types)

    return f"{brand} {adjective} {product}"

def generate_product_description():
    features = [
        "sleek design", "lightweight build", "intuitive functionality", "durable materials",
        "long-lasting battery", "high performance", "smart features", "compact size",
        "ergonomic style", "modern aesthetics"
    ]

    use_cases = [
        "Perfect for home or office use.",
        "Ideal for everyday tasks and travel.",
        "Designed for comfort and efficiency.",
        "A reliable choice for tech lovers.",
        "Built to enhance your lifestyle.",
    ]

    feature1 = random.choice(features)
    feature2 = random.choice(features)
    use_case = random.choice(use_cases)

    description = (
        f"This product combines {feature1} with {feature2}. "
        f"{use_case}"
    )

    return description

def random_decimal(min_value, max_value, decimals=2):
    """
    Genera un número decimal aleatorio entre min_value y max_value
    redondeado a 'decimals' cifras decimales.

    :param min_value: mínimo valor (float)
    :param max_value: máximo valor (float)
    :param decimals: cantidad de decimales para redondear (int, default=2)
    :return: número decimal aleatorio (float)
    """
    number = random.uniform(min_value, max_value)
    return round(number, decimals)

# Random id
def random_index():
    indices = [5,6,7,8,9,10,11,12,13,14]
    return random.choice(indices)

def random_image():
    return random.choice(img_list)

# Return a predifined number of maximum values
def return_max_numbers(n, list_numbers):
    final_list = []
    for i in range(n):
        max_currently_number = max(list_numbers)
        final_list.append(max_currently_number)
        list_numbers.remove(max_currently_number)
    
    return final_list

# Create new country instances based in a range with init and limit
def post_new_countries(init, limit):
    url = "https://restcountries.com/v3.1/all"
    params = {
        "fields": "name,capital,region,population,flags,subregion,area,latlng,translations"
    }

    response = requests.get(url, params=params)

    data = response.json()
    for country in data[init:limit]:
        try:
            Country.objects.create(name=country['name']['common'], region=country['region'], subregion=country['subregion'],
                            capital=country.get('capital', ['Desconocida'])[0], population=country['population'],
                            latitude=country['latlng'][0],  longitude=country['latlng'][1], area=country['area'],
                            flag=country['flags']['alt'])
        except IndexError:
            Country.objects.create(name=country['name']['common'], region=country['region'], subregion=country['subregion'],
                            population=country['population'], latitude=country['latlng'][0],  longitude=country['latlng'][1], 
                            area=country['area'], flag=country['flags']['alt'])
    return response.status_code

# Excel template functions

# Apply border for a cell in a worksheet
def apply_border_to_range(ws, cell_range, border):
    """
    Aplica un borde a todas las celdas dentro del rango dado.
    """
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border

# Apply styles for a specific cell
def apply_cell_styles(cell, styles):
    cell.alignment = styles['alignment']
    cell.font = styles['font']
    cell.fill = styles['fill']

# Insert a personalized text in a worksheet
def insert_text(ws, start_cell, end_cell, text, style=None):
    # Join cells
    cell_range = f"{start_cell}:{end_cell}"
    ws.merge_cells(cell_range)
    cell = ws[start_cell]
    cell.value = text
    if style:
        apply_cell_styles(cell, style)
        apply_border_to_range(ws, cell_range, style['border'])

# Create a excel table using dictionaries with values and styles
def create_excel_table(ws, datos, styles=None, inicio=(1, 1), not_styles=[]):
    # Pull your key and value lists
    encabezados = list(datos.keys())
    filas = zip(*datos.values())  # The rows are the value columns from dictionary

    # Write your initial coordinates
    fila_inicio, col_inicio = inicio

    ws.row_dimensions[fila_inicio].height = 20

    # Write your headers
    for col, encabezado in enumerate(encabezados, start=col_inicio):
        celda = ws.cell(row=fila_inicio, column=col, value=encabezado)
        ws.column_dimensions[get_column_letter(col)].width = len(str(celda.value)) + 1
        # Aplicar estilo de encabezado si se proporciona
        if styles and col not in not_styles:
            apply_cell_styles(celda, styles['headers'])
            celda.border = BORDER

    # Write your main data
    for fila_idx, fila in enumerate(filas, start=fila_inicio + 1):
        for col_idx, valor in enumerate(fila, start=col_inicio):
            ws.row_dimensions[fila_idx].height = 25
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            if ws.column_dimensions[get_column_letter(col_idx)].width < len(str(celda.value)):
                ws.column_dimensions[get_column_letter(col_idx)].width = len(str(celda.value)) + 1
            # Aplicar estilo a las celdas de datos si se proporciona
            if styles and col_idx not in not_styles:
                apply_cell_styles(celda, styles['data'])
                celda.border = BORDER

# matplotlib section

# Apply a new image for your excel report specifying a position and a worksheet
def set_img(ws, position, fig):
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
    plt.close(fig)
    img_buffer.seek(0)

    excel_image = OpenpyxlImage(img_buffer)
    ws.add_image(excel_image, position)

# Apply graph styles for your matplotlib picture/figure
def apply_graph_styles(ws, ax, main_dic, fig, grid=False):
    ax.set_title(main_dic['main_title'], fontsize=main_dic['font_size'], fontweight=main_dic['fontweight'])
    if 'x_title' in main_dic:
        ax.set_xlabel(main_dic['x_title'], fontsize=main_dic['font_size'])
    if 'y_title' in main_dic:
        ax.set_ylabel(main_dic['y_title'], fontsize=main_dic['font_size'])
    ax.tick_params(axis='x', rotation=25)

    fig.patch.set_facecolor(main_dic['facecolor'])  # Fondo exterior
    ax.set_facecolor(main_dic['background'])        # Fondo del área de gráfico
    # Agregar borde alrededor de la figura completa
    fig.patch.set_linewidth(main_dic['border_linewidth'])  # Grosor del borde
    fig.patch.set_edgecolor(main_dic['border_color'])  # Color del borde

    if grid == True:
        ax.grid(color = main_dic['color'], linestyle = main_dic['linestyle'], linewidth = main_dic['linewidth'])
    
    set_img(ws, main_dic['position'], fig)

# Specific function to create a bar graph
def bar(ws, main_dic):
    fig, ax = plt.subplots(figsize=main_dic['figsize'])

    x_labels = main_dic['labels']
    graph_values = main_dic['graph_values']

    # El parámetro edgecolor sirve para darle un color al borde de cada barra
    # bbox_to_anchor sirve para definir el tamaño general del gráfico entero o el ancho
    ax.bar(x_labels, graph_values, label=x_labels, color=main_dic['colors'], edgecolor=main_dic['edgecolor'])
    'Suported values for loc atributte'
    'best', 'upper right', 'upper left', 'lower left', 'lower right', 'right', 'center left', 'center right', 'lower center', 'upper center', 'center'
    ax.legend(title=main_dic['legend_title'], loc="best", bbox_to_anchor=main_dic['bbox_to_anchor'])

    # Show values above each bar
    for i, val in enumerate(graph_values):
        ax.text(i, val+val*.1, f"{val}", ha=main_dic['ha'], fontsize=main_dic['font_size'])

    ax.margins(y=main_dic['margins_y'])

    apply_graph_styles(ws, ax, main_dic, fig)

# Specific function to create a lineal graph
def line(ws, main_dic):
    fig, ax = plt.subplots(figsize=main_dic['figsize'])
    labels = main_dic['labels']
    graph_values = main_dic['graph_values']
    ax.plot(labels, graph_values, marker=main_dic['marker'], linestyle=main_dic['linestyle'], color=main_dic['plot_color'])

    for i, valor in enumerate(graph_values):
        # El parámetro ha sirve para centrar los textos de los valores que se muestran por cada punto
        ax.text(labels[i], valor, str(valor), ha=main_dic['ha'], fontsize=main_dic['font_size'], color=main_dic['plot_text_color'])

    apply_graph_styles(ws, ax, main_dic, fig, True)

# Specific function to create a pie chart
def pie(ws, main_dic):
    fig, ax = plt.subplots()
    # '%1.1f%%' significa:
    # 1.1f: un decimal (por ejemplo, 25.3%)
    # %%: el símbolo de porcentaje
    ax.pie(main_dic['graph_values'], labels=main_dic['labels'], autopct=main_dic['autopct'], shadow=main_dic['shadow'],
        textprops={'color': main_dic['color'], 'font': main_dic['font'], 'weight': main_dic['weight'], 'size': main_dic['size']})
    plt.legend(title=main_dic['legend_title'], loc="center left", bbox_to_anchor=main_dic['bbox_to_anchor'])

    apply_graph_styles(ws, ax, main_dic, fig)

# Specific function to create a histogram
def histogram(ws, main_dic):
    fig, ax = plt.subplots(figsize=main_dic['figsize'])
    graph_values = main_dic['graph_values']
    set_values = set(graph_values)

    n, bins, patches = ax.hist(graph_values, bins=list(set_values), color=main_dic['color'], edgecolor=main_dic['edgecolor'], linestyle=main_dic['linestyle'])

    # Show values for each bar
    for count, bin_left, bin_right in zip(n, bins[:-1], bins[1:]):
        x = (bin_left + bin_right) / 2  # punto medio del bin
        ax.text(x, count+count*.1, str(int(count)), ha=main_dic['ha'], fontsize=main_dic['font_size'])

    ax.margins(y=main_dic['margins_y'])

    apply_graph_styles(ws, ax, main_dic, fig)

# Specific function to create a scatter graph
def scatter(ws, main_dic):
    fig, ax = plt.subplots()
    graph_values = main_dic['graph_values']
    colors = [i for i in range(len(graph_values))]
    graph_values_b = main_dic['graph_values_b']
    sc = ax.scatter(np.array(graph_values), np.array(graph_values_b), c=np.array(colors), cmap='viridis')

    # Show values for every little point
    if 'values' in main_dic:
        for i in range(len(graph_values)):
            ax.text(graph_values[i], graph_values_b[i] + .4, f"({graph_values[i]}, {graph_values_b[i]})", fontsize=9, ha='right', color='black')

    plt.colorbar(sc, ax=ax)

    apply_graph_styles(ws, ax, main_dic, fig, True)